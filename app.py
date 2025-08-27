from flask import Flask, render_template, request 
import openpyxl as op
import pandas as pd
import matplotlib
matplotlib.use('Agg')  # for resolving error (by gpt)
import matplotlib.pyplot as mp

import os
import operator # for string to actual operator
import time    

app = Flask(__name__)

@app.route('/')
def home():
    return render_template('home.html')

@app.route('/page1', methods=["GET", "POST"])
def page1 ():
    graph_url = None
    error = None
    address=""
    need=""
    files=""
   
    
    if request.method == 'POST':
        
        
            final_df = pd.DataFrame()  # created a dataframe 1

            final_df_list = []  # created list of dataframe

            filtered_df = pd.DataFrame()  # created a dataframe 2

            files = request.files.getlist('excelfiles')  # stores list of files 
            
            address = request.form['address']  # stores address 

            a = []   # empty list

            a = [word.strip() for word in address.split('.') if word.strip()]  # list of address words

            sheet_name = a[0].strip()  # sheet name stored (string)!!
            table_name = a[1].strip()  # table name stored (string)!!
            
            need = request.form['need']  # stores requirement or condition    
             
            b = []  # empty list
            b = [need.split('Where')[0].strip(), need.split('Where')[-1].strip()]  # stores required column and condition in list 
            req_col = b[0].strip()  # store required column in string !!!!
            cond = b[1].strip() # store whole condition as string
            cond_list = []   # empty list
            cond_list = [word.strip() for word in cond.split('|') if word.strip()]  # store list of cond. after Where
            no_of_cond = len(cond_list) # checks number of conditions or line on graph
            operand1 = []  # list for storing operands1 in each condition
            operators = [] # list for storing operator in each condition
            operand2 = []  # list for storing operands2 in each condition
            cond_list_labels = []
            for i in range(0,no_of_cond): # iterate to store operand and operator in specific list
                temp_str = cond_list[i]
                temp_str_list = []
                temp_str_list =[word.strip() for word in temp_str.split(" ") if word.strip()]
                operand1.append(temp_str_list[0].strip())   
                operators.append(temp_str_list[1].strip())
                cond_list_labels.append(f"condition {i+1}") #confitions lables list
                if temp_str_list[2].strip().isdigit():
                    xyz = int(temp_str_list[2].strip())
                else:
                    xyz = temp_str_list[2].strip()
                operand2.append(xyz)
            
            
              

            opr = {  # for mapping string operator with actual operator
                "<": operator.lt,
                "<=": operator.le,
                ">": operator.gt,
                ">=": operator.ge,
                "==": operator.eq,
                "!=": operator.ne
            }

            
        
            for file in files:  # iterate in files
                df = pd.DataFrame()  # create new dataframe
                wb = op.load_workbook(file)  # whole workbook
                if sheet_name in wb.sheetnames:  # checks sheet exit in workbook
                    ws = wb[sheet_name]  # opens that sheet

                    
                    for table_obj in ws.tables.values():  # iterate in tables in a sheet
                        if table_obj.name == table_name:  # checks table exist in sheet or not by name
                            
                            data_range = ws[table_obj.ref]  # written tuple 

                            header = [cell.value for cell in data_range[0]]  # list of header columns
                            data = [[cell.value for cell in row] for row in data_range[1:]]  # list of rows(list),containing list of cell values each!!
                            df = pd.DataFrame(data, columns=header)  # creates dataframe of specific table
                            df['src_file'] = file.filename  # creates a column for storing file name in each row
                            filtered_df = pd.concat([filtered_df, df], ignore_index=True)  # store table dataframe from each file in it
                          

            if not filtered_df.empty: # checks df is not empty
                for i in range(0,no_of_cond): # iterate to create dataframe for each conditions
                    func = opr[operators[i]]
                    if operand1[i] in filtered_df.columns: # checks column present in dataframe
                        final_df = filtered_df[func(filtered_df[operand1[i]], operand2[i])].sort_values(by='src_file')  # after applying condition gives final dataframe  
                        final_df_list.append(final_df)  # appends dataframe into list
                    

                

            if final_df_list: #checks list is created
                fig = mp.figure(figsize=(7, 6))  # created new page for graph
                colors =['red','green','blue','balck','brown','pink','yellow','orange']
                for i in range (0,no_of_cond): # plots multiple graph on single graph page
                    mp.plot(final_df_list[i]['src_file'], final_df_list[i][req_col],marker='o',color=colors[i], label = cond_list_labels[i])  # plots points on graph 
                    mp.legend()
                
                mp.xlabel("variation")  # x-axis label
                mp.ylabel(req_col)  # y-axis label
                mp.title(f"Graph of {req_col}")  # title of graph
                graph_file = "graph.png" 
                graph_path = os.path.join(app.root_path, "static", graph_file)  # creates path of image generated
                print("Saving graph to:", graph_path)
                mp.savefig(graph_path)  # saves the graph image generated to specific location or file
                mp.tight_layout()
                mp.grid(True,linestyle='--',alpha=0.5)
                
                mp.close(fig) 
            
                timestamp = int(time.time())  # gpt
                graph_url = f"static/{graph_file}?v={timestamp}"  # gpt
                # graph_url = f"static/{graph_file}"
        
          

    return render_template("page1.html", graph_url=graph_url, message=error,address=address,need=need,files=files)  # passes the data to html webpage after processing (also have input box data)



@app.route('/page2')
def page2():
    return render_template('page2.html')

if __name__ == '__main__':  # check whether script or python app.py is run directly 
    app.run(debug=True, host='0.0.0.0', port=5000)
